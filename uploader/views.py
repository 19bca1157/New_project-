from django.shortcuts import render, redirect, get_object_or_404
from .forms import ExcelFileForm
from .models import ExcelFile
import openpyxl
from datetime import datetime

def home(request):
    date_filter = request.GET.get('date', '')
    files = ExcelFile.objects.all()
    if date_filter:
        files = files.filter(uploaded_at__date=datetime.strptime(date_filter, '%Y-%m-%d').date())

    file_sheets_data = {}
    for file in files:
        wb = openpyxl.load_workbook(file.file.path)
        sheets = {
            sheet: [
                [cell.value for cell in row]
                for row in wb[sheet].iter_rows()
            ]
            for sheet in wb.sheetnames
        }
        file_sheets_data[file] = sheets

    context = {'files': file_sheets_data, 'date_filter': date_filter}
    return render(request, 'uploader/home.html', context)

def upload_file(request):
    if request.method == 'POST':
        form = ExcelFileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = ExcelFileForm()

    return render(request, 'uploader/upload.html', {'form': form})

def view_file(request, file_id):
    file_obj = get_object_or_404(ExcelFile, id=file_id)
    wb = openpyxl.load_workbook(file_obj.file.path)
    sheets = {
        sheet: [
            [cell.value for cell in row]
            for row in wb[sheet].iter_rows()
        ]
        for sheet in wb.sheetnames
    }

    context = {'file': file_obj, 'sheets': sheets}
    return render(request, 'uploader/view.html', context)

def delete_file(request, file_id):
    file_obj = get_object_or_404(ExcelFile, id=file_id)
    file_obj.file.delete(save=False)
    file_obj.delete()
    return redirect('home')

